import shutil
import os
import openpyxl as pyxl
import time
import datetime


class Timesheet_Entry(object):
    def __init__(self, date, wbs_element, start_time):
        self.date = date
        self.wbs_element = wbs_element
        self.start_time = time_entry

    def get_wbs_element(self) 
        return self.wbs_element

    def get_time_entry(self, current_time):
        return time.time() - self.start_time


#initialize time entry class

date = datetime.datetime.today().date()
start_time = time.time()
entry = Timesheet_Entry(date,wbs_element, start_time)
#create timesheet if non-existant
if not os.path.isfile(timesheetname):
    wb=pyxl.Workbook()
    wb.save(Name of timesheet file)



def get_wbs():
    with open txt file:
        if wbs in txt:
            wbs_element = wbs
        else:
            add_wbs #GUI


def new_timesheet():
    #log elapsed project time to timesheet
    log_timesheet_entry
    #send current timesheet to archive if exists
    try:
        os.mkdir(Archive folder)
    except OSError as e:
        print('Archive exists')
        pass
    archive_name = timesheet + current time
    os.rename(timesheet, archive_name)
    shutil.move(archive name, archive folder)
    #create new timesheet excel file
    wb=pyxl.Workbook()
    wb.save(Name of timesheet file)
    #initialize new timesheet entry
    reset_project_timer()
    
def reset_project_timer():
    #sets globabl variable project_start_time to current time
    global project_start_time
    project_start_time = time.time()

def get_elapsed_project_time():
    #returns elapsed time on project
    elapsed_time = time.time()- project_start_time
    return elapsed_time

def log_elapsed_project_time(elapsed_time):
    #read WBS of current InventorWork
    with open txt file:
        network_activity = readlines(second line)
    #initialize default row number var to insert log entry
    wbs_row = None
    #check if wbs already in timesheet, if so, overwrite wbs_row var
    wb = pyxl.load_workbook( timesheet path)
    sheet = wb.get_sheet_by_name('Sheet1')
    for col in sheet.iter_cols(min_row=1, max_col=1, max_row = sheet.max_row):
        for cell in col:
            if network_activity == cell:
                wbs_row = cell.row
                break
            else:
                continue
    #if wbs entry doesn't exist, insert into lowest row
    if wbs_row is None:
        sheet.cell(row=sheet.max_row + 1, column=1, value=
    #check if date entry already exists in timesheet
    if sheet.cell(row=1, column=sheet.max_column).value.date() == datetime.datetime.today().date():
        cell_column = sheet.max_column
    else:
        cell_column = sheet.max_column + 1
    #insert new log entry
    sheet.cell(row=cell_row, column=cell_column, value=
                
        
    
    
    
    
