import subprocess       #To open/close external programs
import time             #To use sleep function, program waits
import shutil           #To rename files/folders
import os               #To search for file names
import psutil           #To search for inventor pid to kill
import win32gui
import win32con
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import datetime
import openpyxl as pyxl


##GUI
##TODO: refactor rename folder to one function


def change(folder):
    #check for folder_name.txt in source folder
    source_path = 'C:\\InventorWork'
    destination_path = 'C:\\' + folder
    if os.path.isfile(source_path + '\\' + 'folder_name.txt'):
        with open(f"{source_path}\\folder_name.txt", 'r') as txtfile:
            source_name = txtfile.read()
    else:
        print('no txt file')
        raise LookupError('Cannot find folder_name.txt')
    #check for folder_name in destination folder
    if os.path.isfile(destination_path + '\\' + 'folder_name.txt'):
        pass
    else:
        no_proj(None, None, folder)     #create folder_name.txt with folder as name
    iw_change = 'C:\\' + folder
    if os.path.exists(iw_change):
        if kill():
            print('closed Inventor')
            try:
                rename()
            except RuntimeError as e:
                raise
            print('ran rename()')
            time.sleep(1)
            print(iw_change)
            time.sleep(1)
            rename_success = False
            for i in range(10):
                try:
                    os.rename(iw_change, 'C:\\InventorWork')
                    rename_success = True
                    break
                except PermissionError:
                    time.sleep(1)
                    continue
            if not rename_success:
                #rename folder back to previous inventorwork
                try:
                    print(f'source name C:\\{source_name}')
                    os.rename(f'C:\\{source_name}', 'C:\\InventorWork')
                    raise RuntimeError
                except PermissionError:
                    print
                    raise RuntimeError
            print('renamed from' + iw_change)
            open_inv()
            print('opened inventor')
            print('done')
            return True
        else:
            print('inventor not closed')
            raise OSError('Inventor not closed')

def make(project,machine):
    if kill():
        try:
            rename()
            newIW(project, machine)
        except RuntimeError as e:
            raise
    else:
        print('inventor not closed')
        raise OSError('Inventor not closed')
    

##Close current inventor session

def kill():
    handle = WindowEnumerate()
    print(handle)
    if not handle:
        print('inventor already closed')
        return True
    win32gui.SetForegroundWindow(handle)
    win32gui.SendMessage(handle,win32con.WM_CLOSE,0,0)  #waits for message to be processed
    if not WindowEnumerate():
        print('kill returning True')
        return True
    else:
        print('kill returning False')
        return False

#windows enumeration handler functions, enumerates windows to get handlers
    #then gets window names
    #adapted from: https://www.blog.pythonlibrary.org/2014/10/20/pywin32-how-to-bring-a-window-to-front/

def windowEnumerationHandler(hwnd, top_windows):
    top_windows.append((hwnd, win32gui.GetWindowText(hwnd)))

def WindowEnumerate():
    top_windows = []
    win32gui.EnumWindows(windowEnumerationHandler, top_windows)
    for i in top_windows:
        if 'Autodesk Inventor Professional' in i[1]:
            print(i)
            return i[0]
    return False
    

def open_inv():
    p = subprocess.Popen('C:\\Program Files\\Autodesk\\Inventor 2016\\Bin\\Inventor.exe');
    print('opened inventor')
    print(p.poll() == None);
    time.sleep(5);
    print(p.poll() == None);



#Rename current inventorwork folder to match folder_name.txt
def rename():
    txtpath = 'C:\\InventorWork\\folder_name.txt'
    folder_txt = open(txtpath)
    source_name = folder_txt.read()
    folder_txt.close()
    time.sleep(1)
    result = False
    current_time = time.time()
    while (time.time() - current_time) < 30:
        try:
            os.rename('C:\\InventorWork', 'C:\\' + source_name)
            print('Renamed InventorWork to ' + source_name)
            return True
        except PermissionError:
            time.sleep(0.5)
            print(time.time() - current_time)
            continue
    raise RuntimeError


def newIW(project, machine):
    p_name = 'InventorWork_' + project.replace(' ', '') + '_' + machine.replace(' ', '') # strips input spaces formates to _project_machine
    batf = 'M:\\Project Engineering\\ENGDATA\\Reference Data\\Dev and Training Resc\\cad\\30 Inventor\\New InventorWork.bat'
    os.startfile(batf)         #inventor must be opened to generate project file in IW
    time.sleep(2)       #sleep buffer to make sure project file is generated, not sure if required?
    no_proj(project,machine, None)
    open_inv()
    return 'done newIW'

def no_proj(project,machine, folder):
    #if folder name input, generate folder_name.txt file in folder input
    if folder != None:
        destination_path = 'C:\\' + folder
        open(destination_path + '\\folder_name.txt', 'a+')
        folder_txt = open(destination_path + '\\folder_name.txt', 'w')
        folder_txt.write(folder)
        folder_txt.close
    #else generate txt from project and machine input
    else:
        open('C:\\InventorWork\\folder_name.txt', 'a+')
        folder_txt = open('C:\\InventorWork\\folder_name.txt', 'w')
        folder_txt.write('InventorWork_' + project + '_' + machine)
        folder_txt.close
    
def check_valid_folder_name(name):
    for i in name:
        if ord(i) > 64 and ord(i) < 91:
            continue
        if ord(i) > 96 and ord(i) < 123:
            continue
        if ord(i) > 47 and ord(i) < 58:
            continue
        if ord(i) == 95 or ord(i) == 32:
            continue
        else:
            return False
    return True
'''
def rename_folder(project: str, machine:str, folder:str) -> None:
    with open(f"C:\\{folder}\\folder_name.txt", 'w+') as txtfile:
        if project and machine is None:
            
        txtfile.write(f"InventorWork_{project}_{machine}")
    if folder != 'InventorWork':
        try:
            os.rename(f"C:\\{folder}", f"C:\\InventorWork_{project}_{machine}")
    else:
        pass

'''

### Time sheet functions ########


class Timesheet_Entry:

    def __init__(self, date, start_time, wbs):
        self.date = date
        self.start_time = start_time
        self.wbs = wbs

    def get_date(self):
        return self.date

    def get_time(self):
        return (time.time() - self.start_time)/3600

    def get_wbs(self):
        return self.wbs


def log_timesheet_entry(entry :Timesheet_Entry):
    #takes valid timesheet entry object and logs into current active
    #timesheet
    try:
        wb = pyxl.load_workbook('C:\\Users\\active_timesheet.xlsx')
        sheet= wb.active
    except NameError:
        raise
    #set default timesheet entry row number as max+1
    entry_row = sheet.max_row + 1
    #check current wbs entries, overwrite row entry if existing in sheet
    for row in sheet.iter_rows(min_row=1, max_col=1,max_row=sheet.max_row):
        for cell in row:
            try:
                if entry.get_wbs() in cell.value:
                    entry_row = cell.row
                    break
                else:
                    continue
                break
            except TypeError:
                continue
    # if wbs entry doesn't exist create
    if entry_row == (sheet.max_row + 1):
        sheet.cell(row = entry_row, column = 1).value = entry.get_wbs()
    #check date entries, max column entry should be latest
    latest_date_entry = sheet.cell(row=1, column=sheet.max_column).value
    if type(latest_date_entry) == datetime.datetime:
        if entry.get_date() == latest_date_entry.date():
            entry_column = sheet.max_column
        else:
            entry_column = sheet.max_column + 1
            sheet.cell(row=1, column = entry_column).value = entry.get_date()
    elif latest_date_entry is None:
        entry_column = sheet.max_column + 1
        sheet.cell(row=1, column = entry_column).value = entry.get_date()
    else:
        raise TypeError
    #check for existing time entry
    entry_cell = sheet.cell(row = entry_row, column = entry_column)
    #add time entry to existing entry if it exists
    if entry_cell.value is not None:
        entry_cell.value += entry.get_time()
    else:
        entry_cell.value = entry.get_time()
    wb.save('C:\\Users\\active_timesheet.xlsx')

def ask_log_time_entry():
    wbs = active_timesheet_entry.get_wbs()
    time = active_timesheet_entry.get_time()
    if messagebox.askyesno('Log Timesheet Entry', f'Are you sure you want to log {time} hours to {wbs} in your timesheet?'):
        log_timesheet_entry(active_timesheet_entry)
        root.destroy()
    else:
        root.destroy()

def save_wbs_data(wbs):
    with open('C:\\InventorWork\\wbs.txt', 'w+') as txtfile:
        txtfile.write(wbs)
    return

def instantiate_timesheet_entry(wbs=None) -> Timesheet_Entry:
    if wbs is None:
        with open('C:\\InventorWork\\wbs.txt', 'r') as file:
            wbs = file.read()
    date = datetime.datetime.now().date()
    start_time = time.time()
    return Timesheet_Entry(date, start_time, wbs)
    
def check_for_wbs_data():
    if os.path.isfile('C:\\InventorWork\\wbs.txt'):
        return True
    else:
        return False
    
    
### GUI ################

class igui:

    def __init__(self, master):
        self.master = master            #master = root = tk()
        self.master.title("InventorWork Manager")   #master.title = root.title = tk().title
        
        #initializing tkk frame to hold widgets, sets size to expand and padding
        self.mainframe = ttk.Frame(master, padding = '3 3 12 12')
        self.mainframe.grid(column = 0, row = 0, sticky = (N,S,E,W))
        self.mainframe.columnconfigure(0, weight = 1)
        self.mainframe.rowconfigure(0, weight = 1)

        #labels and buttons
        self.newlabel = ttk.Label(self.mainframe, text = 'New InventorWork',
                                  justify = 'left')
        self.newlabel.grid(column = 1, row = 1, sticky = W)
        self.changelabel = ttk.Label(self.mainframe, text = 'Change InventorWork',
                                     justify = 'left')
        self.changelabel.grid(column = 1, row = 4, sticky = W)
        self.NewButton = ttk.Button(self.mainframe, text = 'New InventorWork', command = self.callwindow)
        self.NewButton.grid(column = 1, row = 2, sticky = W)
        self.ChangeButton = ttk.Button(self.mainframe, text = 'Change InventorWork', command = self.yousure)
        self.ChangeButton.grid(column = 2, row = 5, sticky = (E,W))
        self.RenameButton = ttk.Button(self.mainframe, text = 'Rename InventorWork', command = self.callrename)
        self.RenameButton.grid(column = 3, row = 5, sticky = (E,W))
        self.HelpButton = ttk.Button(self.mainframe, text='?', command = self.help)
        self.HelpButton.grid(column = 3, row = 1, sticky = (W))
        #scrollbar for folder selection and seperator
        self.folder_var = StringVar()
        self.separator = ttk.Separator(self.mainframe, orient = HORIZONTAL).grid(row = 3, columnspan = 4, sticky = (E, W))
        self.iscroll = ttk.Combobox(self.mainframe, textvariable = self.folder_var,
                                    width = 40)

        
        self.iscroll.bind('<<ComboboxSelected>>', self.comboselect)
        self.iscroll.grid(column = 1, row = 5)
        self.current_folder_name = StringVar()
        
        self.current_folder_label = ttk.Label(self.mainframe, textvariable = self.current_folder_name,
                                     justify = 'left')
        self.current_folder_label.grid(column = 1, row = 6, sticky = W)
        self.update_list()
        for child in self.mainframe.winfo_children():
            child.grid_configure(padx=5, pady=5)

        self.start_timesheet_entry()



    #function called everyime scroll window selection changed, sets variable
            #equal to selection
    def comboselect(self,eventObject):
        self.change_selection = self.folder_var.get()
        print(self.change_selection)
        return self.change_selection

    #function called when change button is clicked

    def changeclick(self):
        try:
            change(self.change_selection)
            self.update_list()
            print('folder changed successfully')
            self.start_time_sheet_entry()
        except LookupError as e:
            print(repr(e))
            self.rename_load.destroy()
            self.rename_load.update()
            self.noproject = messagebox.showinfo(message = "No folder_name.txt file in InventorWork folder. Please name the"
                                                             " current InventorWork folder.")
            self.cwindow = Toplevel(self.master)
            self.app = window_change(self.cwindow)
            self.cwindow.grab_set()
            self.master.wait_window(self.cwindow)
            print('load window SHOULD BE HERE')
            self.rename_load = Toplevel(self.master)
            self.rename_load.transient()
            self.app2 = loadwindow(self.rename_load)
            self.rename_load.update()
            if self.app.changepressed:
                change(self.change_selection)
            print('from changeclick func')
            self.rename_load.destroy()
            self.rename_load.update()
            self.update_list()
            self.start_timesheet_entry()
        except OSError as e:
            print(repr(e))
            self.update_list()
        except RuntimeError as e:
            print(repr(e))
            self.rename_load.destroy()
            self.rename_load.update()
            messagebox.showinfo(message = "Renaming folder timed out. Please ensure no programs are using the InventorWork folder.")
            self.update_list()



        

    def yousure(self):
        #check there is a folder selected
        try:
            print(self.change_selection)
            self.yousure = messagebox.askyesno('Restart', 'Are you sure you want to restart Inventor and change InventorWork folder?', parent = self.master)
            if self.yousure:
                self.rename_load = Toplevel(self.master)
                self.rename_load.transient()
                self.app = loadwindow(self.rename_load)
                self.rename_load.update()
                print ('supposed to have ran loading window')
                self.changeclick()
                self.rename_load.destroy()
            else:
                pass
        except AttributeError as e:
            print(f"No folder selected, {e}")
            pass


    def help(self):
        self.helpmessage = messagebox.showinfo(message = "Version 2.0 \n\nThis app manages InventorWork folders. InventorWork names are kept"
                                                           " track of by generating a folder_name.txt file in the InventorWork"
                                                           " folder. The user is prompted if a folder_name.txt file"
                                                            " doesn't exist. Inventor is restarted when changing folders if Inventor is open.\n\n Written by Adrian Tai")
        
            

    #method creates new window and passes new window as master into class that
            #contains the labels and buttons for change IW window
    def callwindow(self):
        if os.path.isfile('C:\\InventorWork\\folder_name.txt'):
            self.mwindow = Toplevel(self.master)
            results = window_make_txt_exists(self.mwindow)
            print('updating list')
            self.update_list()
        else:
            self.mwindow = Toplevel(self.master)
            self.app = window_make_notxt(self.mwindow)
            self.mwindow.grab_set()
            self.master.wait_window(self.mwindow)
            self.update_list()
            



    def update_list(self):
        print('list update')
        self.folder_list =[]
        for i in os.listdir('C:\\'):
            if 'InventorWork' in i:
                self.folder_list.append(i)
        self.iscroll['values'] = self.folder_list
        txtpath = 'C:\\InventorWork\\folder_name.txt'
        current_folder_name = ''
        try:
            with open(txtpath) as file:
                current_folder = file.read().split('_')
                for i in range(1, len(current_folder)):
                    current_folder_name += (current_folder[i] + ' ')
        except FileNotFoundError:
            pass
        print(f"current folder name: {current_folder_name}")
        self.current_folder_name.set('Current InventorWork: ' + current_folder_name)


    def callrename(self):
        try:
            print(self.change_selection)
            self.rwindow = Toplevel(self.master)
            self.app = rename_window(self.rwindow, self.change_selection)
            #self.rwindow.grab_set()
            self.master.wait_window(self.rwindow)
            self.update_list()
        except AttributeError:
            pass


    def start_timesheet_entry(self):
        #check for wbs data ask if unavailable
        if  not check_for_wbs_data:
            messagebox.showinfo(message = "Please enter network and activity for the current machine")
            self.mwindow = Toplevel(self.master)
            wbs_window = Wbs_Entry_Window(self.mwindow)
            print('updating list')
            self.update_list()
        else:
            global active_timesheet_entry
            active_timesheet_entry = instantiate_timesheet_entry()
        

class window_make_txt_exists:
    def __init__(self, master):
        self.master = master     #sets self.master= master = initializing variable input
                                # intializing varialbe = self.mwindow =
                                #Toplevel(self.mainframe) = new window with mainframe
                                #as parent
        
        self.project = StringVar()
        self.machine = StringVar()
        self.master.rowconfigure(1, pad = 20, weight = 1)
        self.master.rowconfigure(4, pad = 20, weight = 1)
        self.main_label = ttk.Label(self.master, text = 'New InventorWork')
        self.main_label.grid(column = 1, row =1, columnspan = 2)
        self.project_label = ttk.Label(self.master, text = 'Project', justify = 'center').grid(column = 1, row = 2, sticky = 'e')
        self.machine_label = ttk.Label(self.master, text = 'Machine', justify = 'center').grid(column = 1, row = 3)
        self.project_entry = ttk.Entry(self.master, textvariable = self.project).grid(column = 2, row = 2)
        self.machine_entry = ttk.Entry(self.master, textvariable = self.machine).grid(column = 2, row =3)
        self.makebutton = ttk.Button(self.master, text = 'Make', command = self.newclick).grid(row = 4, column=2, columnspan = 2)
        for child in self.master.winfo_children():
            child.grid_configure(padx = 10, pady = 5)
        master.grab_set()
        master.wait_window()


    def newclick(self):
        if not (self.project.get() and self.machine.get()):
            return False
        if not (check_valid_folder_name(self.project.get()) and check_valid_folder_name(self.machine.get())):
                messagebox.showinfo(message = "Names can only contain alphanumeric characters.")
        else:
            self.yesno = messagebox.askyesno('Restart', "Are you sure you want to restart Inventor and create new InventorWork?", parent = self.master)
            if self.yesno:
                self.rename_load = Toplevel(self.master)
                self.rename_load.transient()
                self.app = loadwindow(self.rename_load)
                self.rename_load.update()
                print ('supposed to have ran loading window')
                print('destroyed toplevel')
                try:
                    check = make(self.project.get(), self.machine.get())
                    self.start_timesheet_entry()
                except OSError as e:
                    print(repr(e))
                    self.update_list()
                except RuntimeError as e:
                    print(repr(e))
                    self.rename_load.destroy()
                    self.rename_load.update()
                    messagebox.showinfo(message = "Renaming folder timed out. Please ensure no programs are using the InventorWork folder.")
                    self.update_list()
                self.master.destroy()
                if check:
                    return True
                else:
                    return False

class window_change:
    def __init__(self, master):
        self.master = master     #sets self.master= master = initializing variable input
                                # intializing varialbe = self.mwindow =
                                #Toplevel(self.mainframe) = new window with mainframe
                                #as parent
        #variable detetcing ok button click
        self.changepressed = False
        self.project = StringVar()
        self.machine = StringVar()
        self.master.rowconfigure(4, pad = 20, weight = 1)
        self.master.rowconfigure(1, pad = 20, weight = 1)
        self.main_label = ttk.Label(self.master, text = 'No folder_name.txt found for current InventorWork. Please name current InventorWork.',justify = 'center')
        self.main_label.grid(row = 1, column = 1, columnspan = 2)
        self.project_label = ttk.Label(self.master, text = 'Project').grid(column = 1, row = 2, sticky = 'e')
        self.machine_label = ttk.Label(self.master, text = 'Machine').grid(column = 1, row = 3, sticky = 'e')
        self.project_entry = ttk.Entry(self.master, textvariable = self.project).grid(column = 2, row = 2)
        self.machine_entry = ttk.Entry(self.master, textvariable = self.machine).grid(column = 2, row =3)
        self.changebutton = ttk.Button(self.master, text = 'Change', command = self.changeclick2).grid(row = 4, column=2, columnspan = 2)
        for child in self.master.winfo_children():
            child.grid_configure(padx = 5, pady = 5)


    def changeclick2(self):
        if not (self.project.get() and self.machine.get()):
            return False
        if not (check_valid_folder_name(self.project.get()) and check_valid_folder_name(self.machine.get())):
            messagebox.showinfo(message = "Names can only contain alphanumeric characters.")
        else:
            self.yesno = messagebox.askyesno('Restart', "Are you sure you want to restart Inventor and change InventorWork?", parent = self.master)
            if self.yesno:
                no_proj(self.project.get(), self.machine.get(), None)
                self.master.destroy()
                self.changepressed = True
                print('destroyed toplevel')

class window_make_notxt:
    def __init__(self, master):
            
        self.master = master     #sets self.master= master = initializing variable input
                                # intializing varialbe = self.mwindow =
                                #Toplevel(self.mainframe) = new window with mainframe
                                #as parent
        
        self.project_old = StringVar()
        self.machine_old = StringVar()
        self.master.rowconfigure(1, pad = 20, weight = 1)
        self.master.rowconfigure(5, pad = 30, weight = 1)
        self.main_label = ttk.Label(self.master, text = 'No folder_name.txt found, please name the current and new InventorWork folders', justify='center')
        self.main_label.grid(row = 1, column = 1, columnspan = 4)
        self.ciw_label = ttk.Label(self.master, text = 'Current InventorWork', justify = 'center')
        self.ciw_label.grid(row = 2, column = 2)
        self.niw_label = ttk.Label(self.master, text = 'New InventorWork', justify = 'center')
        self.niw_label.grid(row = 2, column = 5) 
        
        self.project_label_old = ttk.Label(self.master, text = 'Project', justify = 'center').grid(column = 1, row = 3)
        self.machine_label_old = ttk.Label(self.master, text = 'Machine', justify = 'center').grid(column = 1, row = 4)
        self.project_entry_old = ttk.Entry(self.master, textvariable = self.project_old).grid(column = 2, row = 3)
        self.machine_entry_old = ttk.Entry(self.master, textvariable = self.machine_old).grid(column = 2, row =4)

        self.separator = ttk.Separator(self.master, orient = VERTICAL).grid(column = 3, row = 2, rowspan = 3, sticky = (N, S))

        self.project_new = StringVar()
        self.machine_new = StringVar()
        self.project_label_new = ttk.Label(self.master, text = 'Project', justify = 'center').grid(column = 4, row = 3)
        self.machine_label_new = ttk.Label(self.master, text = 'Machine', justify = 'center').grid(column = 4, row = 4)
        self.project_entry_new = ttk.Entry(self.master, textvariable = self.project_new).grid(column = 5, row = 3)
        self.machine_entry_new = ttk.Entry(self.master, textvariable = self.machine_new).grid(column = 5, row =4)
        
        self.makebutton = ttk.Button(self.master, text = 'Make', command = self.newclick).grid(row = 5, column=5, columnspan = 4)
        for child in self.master.winfo_children():
            child.grid_configure(padx = 20, pady = 5)

    def newclick(self):
        if not (self.project_new.get() and self.machine_new.get() and self.project_old.get() and self.machine_old.get()):
            return False
        if not (check_valid_folder_name(self.project_new.get()) and check_valid_folder_name(self.machine_new.get()) and check_valid_folder_name(self.project_old.get()) and check_valid_folder_name(self.machine_old.get())):
            messagebox.showinfo(message = "Names can only contain alphanumeric characters.")
        else:
            self.yesno = messagebox.askyesno('Restart', "Are you sure you want to restart Inventor and create new InventorWork?", parent = self.master)
            if self.yesno:
                self.rename_load = Toplevel(self.master)
                self.rename_load.transient()
                self.app = loadwindow(self.rename_load)
                self.rename_load.update()
                print('destroyed toplevel')
                no_proj(self.project_old.get(), self.machine_old.get(), None)
                check = make(self.project_new.get(), self.machine_new.get())
                self.master.destroy()

class loadwindow:
    def __init__(self, master):
        self.master = master
        self.master.attributes("-topmost", True)
        self.mainlabel = ttk.Label(self.master, text = 'Folder may be in use, trying to rename folder...', justify = 'center')
        self.mainlabel.grid_configure(padx = 40, pady = 40)
#rename window -- new
class rename_window:
    def __init__(self, master, folder):
        self.master = master     #sets self.master= master = initializing variable input
                                # intializing varialbe = self.mwindow =
                                #Toplevel(self.mainframe) = new window with mainframe
                                #as parent
        
        self.project = StringVar()
        self.machine = StringVar()
        self.folder = folder #folder variable for rename function
        self.master.rowconfigure(1, pad = 20, weight = 1)
        self.master.rowconfigure(4, pad = 20, weight = 1)
        self.main_label = ttk.Label(self.master, text = 'Rename InventorWork')
        self.main_label.grid(column = 1, row =1, columnspan = 2)
        self.project_label = ttk.Label(self.master, text = 'Project', justify = 'center').grid(column = 1, row = 2, sticky = 'e')
        self.machine_label = ttk.Label(self.master, text = 'Machine', justify = 'center').grid(column = 1, row = 3)
        self.project_entry = ttk.Entry(self.master, textvariable = self.project).grid(column = 2, row = 2)
        self.machine_entry = ttk.Entry(self.master, textvariable = self.machine).grid(column = 2, row =3)
        self.makebutton = ttk.Button(self.master, text = 'Rename', command = self.renameclick).grid(row = 4, column=2, columnspan = 2)
        for child in self.master.winfo_children():
            child.grid_configure(padx = 10, pady = 5)
        master.grab_set()
        #master.wait_window()


    def renameclick(self):
        if not (self.project.get() and self.machine.get()):
            return False
        if not (check_valid_folder_name(self.project.get()) and check_valid_folder_name(self.machine.get())):
                messagebox.showinfo(message = "Names can only contain alphanumeric characters.")
        else:
            self.yesno = messagebox.askyesno('Rename', "Are you sure you want to rename InventorWork?", parent = self.master)
            if self.yesno:
                rename_folder(self.project.get(),self.machine.get(),self.folder)
                self.master.destroy()
                
class Wbs_Entry_Window:
    
    def __init__(self, master):
        self.master = master     #sets self.master= master = initializing variable input
                                # intializing varialbe = self.mwindow =
                                #Toplevel(self.mainframe) = new window with mainframe
                                #as parent
        
        self.network = StringVar()
        self.activity = StringVar()
        self.master.rowconfigure(1, pad = 20, weight = 1)
        self.master.rowconfigure(4, pad = 20, weight = 1)
        self.main_label = ttk.Label(self.master, text = 'WBS Entry')
        self.main_label.grid(column = 1, row =1, columnspan = 2)
        self.network_label = ttk.Label(self.master, text = 'Network', justify = 'center').grid(column = 1, row = 2, sticky = 'e')
        self.activity_label = ttk.Label(self.master, text = 'Activity', justify = 'center').grid(column = 1, row = 3)
        self.network_entry = ttk.Entry(self.master, textvariable = self.network).grid(column = 2, row = 2)
        self.activity_entry = ttk.Entry(self.master, textvariable = self.activity).grid(column = 2, row =3)
        self.makebutton = ttk.Button(self.master, text = 'OK', command = self.set_wbs).grid(row = 4, column=2, columnspan = 2)
        for child in self.master.winfo_children():
            child.grid_configure(padx = 10, pady = 5)
        master.grab_set()
        master.wait_window()

    def set_wbs(self):
        if not (self.network.get() and self.activity.get()):
            return False
        else:
            network = self.network.get()
            activity = self.activity.get()
            wbs = f"{network}/{activity}"
            global active_timesheet_entry
            active_timesheet_entry = instantiate_timesheet_entry(wbs)
            save_wbs_data(wbs)
            self.master.destroy()
        



    
    


        
#create timesheet if non-existant
if not os.path.isfile('C:\\Users\\active_timesheet.xlsx'):
    wb=pyxl.Workbook()
    wb.save('C:\\Users\\active_timesheet.xlsx')
#timesheet entry global var
active_timesheet_entry = None
root = Tk()
app = igui(root)
root.protocol("WM_DELETE_WINDOW", ask_log_time_entry)
root.mainloop()
  
